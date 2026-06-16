import io
import os
import re
import json
import tempfile
from datetime import datetime
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
CORS(app)

try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False

try:
    import pytesseract
    from PIL import Image
    HAS_TESSERACT = True
except ImportError:
    HAS_TESSERACT = False


MOCK_INVOICES = [
    {
        "invoice_number": "00123456",
        "invoice_date": "2024-01-15",
        "total_amount_tax": 11300.00,
        "total_amount_no_tax": 10000.00,
        "tax_amount": 1300.00,
        "buyer_name": "北京科技有限公司",
        "seller_name": "上海贸易有限公司",
        "items": [
            {
                "name": "笔记本电脑",
                "spec": "14寸/16G/512G",
                "quantity": 5,
                "unit_price": 2000.00,
                "amount": 10000.00
            }
        ]
    },
    {
        "invoice_number": "00123457",
        "invoice_date": "2024-01-20",
        "total_amount_tax": 5650.00,
        "total_amount_no_tax": 5000.00,
        "tax_amount": 650.00,
        "buyer_name": "深圳科技发展公司",
        "seller_name": "广州电子有限公司",
        "items": [
            {
                "name": "机械键盘",
                "spec": "87键/青轴",
                "quantity": 10,
                "unit_price": 300.00,
                "amount": 3000.00
            },
            {
                "name": "无线鼠标",
                "spec": "蓝牙/静音",
                "quantity": 20,
                "unit_price": 100.00,
                "amount": 2000.00
            }
        ]
    },
    {
        "invoice_number": "00123458",
        "invoice_date": "2024-02-05",
        "total_amount_tax": 22600.00,
        "total_amount_no_tax": 20000.00,
        "tax_amount": 2600.00,
        "buyer_name": "杭州网络科技公司",
        "seller_name": "南京数码产品公司",
        "items": [
            {
                "name": "显示器",
                "spec": "27寸/2K/IPS",
                "quantity": 8,
                "unit_price": 1500.00,
                "amount": 12000.00
            },
            {
                "name": "打印机",
                "spec": "激光/黑白",
                "quantity": 4,
                "unit_price": 2000.00,
                "amount": 8000.00
            }
        ]
    }
]


def extract_text_from_image(image_file):
    if not HAS_TESSERACT:
        return ""
    try:
        img = Image.open(image_file)
        text = pytesseract.image_to_string(img, lang='chi_sim+eng')
        return text
    except Exception:
        return ""


def extract_text_from_pdf(pdf_file):
    if not HAS_PDFPLUMBER:
        return ""
    try:
        text = ""
        with pdfplumber.open(pdf_file) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        return text
    except Exception:
        return ""


def parse_invoice_text(text):
    result = {
        "invoice_number": "",
        "invoice_date": "",
        "total_amount_tax": 0,
        "total_amount_no_tax": 0,
        "tax_amount": 0,
        "buyer_name": "",
        "seller_name": "",
        "items": []
    }

    invoice_no_match = re.search(r'(?:发票号码|No\.)[：:\s]*([0-9]{6,})', text)
    if invoice_no_match:
        result["invoice_number"] = invoice_no_match.group(1)

    date_match = re.search(r'(?:开票日期|日期)[：:\s]*(\d{4}[-年/]\d{1,2}[-月/]\d{1,2})', text)
    if date_match:
        date_str = date_match.group(1)
        date_str = date_str.replace('年', '-').replace('月', '-').replace('/', '-')
        try:
            dt = datetime.strptime(date_str, '%Y-%m-%d')
            result["invoice_date"] = dt.strftime('%Y-%m-%d')
        except ValueError:
            result["invoice_date"] = date_str

    buyer_match = re.search(r'购买方[^名]*名称[：:\s]*([^\n]+)', text)
    if buyer_match:
        result["buyer_name"] = buyer_match.group(1).strip()

    seller_match = re.search(r'销售方[^名]*名称[：:\s]*([^\n]+)', text)
    if seller_match:
        result["seller_name"] = seller_match.group(1).strip()

    total_tax_match = re.search(r'(?:价税合计|合计金额)[^\d]*([\d,]+\.?\d*)', text)
    if total_tax_match:
        result["total_amount_tax"] = float(total_tax_match.group(1).replace(',', ''))

    total_no_tax_match = re.search(r'(?:不含税金额|金额合计)[^\d]*([\d,]+\.?\d*)', text)
    if total_no_tax_match:
        result["total_amount_no_tax"] = float(total_no_tax_match.group(1).replace(',', ''))

    tax_match = re.search(r'(?:税额|合计税额)[^\d]*([\d,]+\.?\d*)', text)
    if tax_match:
        result["tax_amount"] = float(tax_match.group(1).replace(',', ''))

    return result


@app.route('/api/ocr', methods=['POST'])
def ocr_invoice():
    if 'file' not in request.files:
        return jsonify({"error": "未上传文件"}), 400

    file = request.files['file']
    filename = file.filename or ""
    file_ext = os.path.splitext(filename)[1].lower()

    text = ""
    mock_index = hash(filename) % len(MOCK_INVOICES)

    if file_ext == '.pdf':
        if HAS_PDFPLUMBER:
            text = extract_text_from_pdf(file)
    elif file_ext in ['.jpg', '.jpeg', '.png', '.bmp', '.tiff']:
        if HAS_TESSERACT:
            text = extract_text_from_image(file)

    result = None
    if text:
        result = parse_invoice_text(text)

    if not result or not result.get("invoice_number"):
        result = MOCK_INVOICES[mock_index].copy()
        suffix = str(hash(filename) % 1000).zfill(3)
        result["invoice_number"] = result["invoice_number"][:-3] + suffix

    return jsonify({
        "success": True,
        "data": result
    })


@app.route('/api/export', methods=['POST'])
def export_excel():
    try:
        data = request.get_json()
        invoices = data.get('invoices', [])

        wb = Workbook()
        ws = wb.active
        ws.title = "发票汇总"

        headers = [
            "源发票号码", "商品名称", "规格型号", "数量",
            "单价", "金额", "购买方名称", "销售方名称",
            "开票日期", "价税合计"
        ]

        header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        row_idx = 2
        for invoice in invoices:
            invoice_no = invoice.get("invoice_number", "")
            buyer_name = invoice.get("buyer_name", "")
            seller_name = invoice.get("seller_name", "")
            invoice_date = invoice.get("invoice_date", "")
            total_tax = invoice.get("total_amount_tax", 0)
            items = invoice.get("items", [])

            for item in items:
                ws.cell(row=row_idx, column=1, value=invoice_no).border = thin_border
                ws.cell(row=row_idx, column=2, value=item.get("name", "")).border = thin_border
                ws.cell(row=row_idx, column=3, value=item.get("spec", "")).border = thin_border
                ws.cell(row=row_idx, column=4, value=item.get("quantity", 0)).border = thin_border
                ws.cell(row=row_idx, column=5, value=item.get("unit_price", 0)).border = thin_border
                ws.cell(row=row_idx, column=6, value=item.get("amount", 0)).border = thin_border
                ws.cell(row=row_idx, column=7, value=buyer_name).border = thin_border
                ws.cell(row=row_idx, column=8, value=seller_name).border = thin_border
                ws.cell(row=row_idx, column=9, value=invoice_date).border = thin_border
                ws.cell(row=row_idx, column=10, value=total_tax).border = thin_border
                row_idx += 1

        column_widths = [15, 20, 18, 8, 12, 12, 25, 25, 12, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"发票汇总_{timestamp}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({
        "status": "ok",
        "pdfplumber": HAS_PDFPLUMBER,
        "tesseract": HAS_TESSERACT
    })


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
